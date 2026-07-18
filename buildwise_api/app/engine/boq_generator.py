"""
BuildWise AI — BOQ Generator (Stage 11)
==========================================
Generates structured Bill of Quantities at three levels:
  1. Room-wise BOQ — Materials for each individual room
  2. Floor-wise BOQ — Aggregated by floor
  3. Building BOQ — Total project summary

Supports export to PDF (ReportLab), Excel (openpyxl), and CSV.
"""

import math
import io
import csv
from typing import List, Dict, Any, Optional
from datetime import datetime

from app.engine.is_codes import (
    BRICKS_PER_M3, AAC_BLOCKS_PER_M3,
    MORTAR_FRACTION_OF_MASONRY, MORTAR_DRY_VOLUME_FACTOR,
    MORTAR_MIX, PLASTER_THICKNESS_MM, CEMENT_BAG_VOLUME_M3,
    CONCRETE_DRY_VOLUME_FACTOR, CONCRETE_MIX_RATIOS,
    CPWD_BASE_RATES_2024,
)


class BOQGenerator:
    """Generate Bill of Quantities from analysis results and material estimates."""

    @staticmethod
    def generate(
        analysis: Dict[str, Any],
        materials: Dict[str, Any],
        costs: Dict[str, Any],
        project_name: str = "BuildWise Project",
        floor_height_m: float = 3.0,
        wall_thickness_m: float = 0.230,
        concrete_grade: str = "M20",
        brick_type: str = "red_brick",
    ) -> Dict[str, Any]:
        """
        Generate comprehensive BOQ from analysis results.

        Returns:
            Dict containing building_boq, room_wise_boq, and metadata.
        """
        rooms = analysis.get("rooms", [])
        rates = dict(CPWD_BASE_RATES_2024)
        if costs and "rates_applied" in costs:
            rates.update(costs["rates_applied"])

        # ── Building-Level BOQ ────────────────────────────────────────────
        building_boq = BOQGenerator._generate_building_boq(
            materials, costs, rates
        )

        # ── Room-wise BOQ ─────────────────────────────────────────────────
        room_wise_boq = BOQGenerator._generate_room_wise_boq(
            rooms=rooms,
            floor_height_m=floor_height_m,
            wall_thickness_m=wall_thickness_m,
            brick_type=brick_type,
            rates=rates,
        )

        return {
            "project_name": project_name,
            "generated_at": datetime.utcnow().isoformat(),
            "building_boq": building_boq,
            "room_wise_boq": room_wise_boq,
            "summary": {
                "total_rooms": len(rooms),
                "building_area_m2": analysis.get("building_area_sq_m", 0),
                "building_area_sqft": analysis.get("building_area_sq_ft", 0),
                "grand_total": costs.get("grand_total", 0),
                "currency": "INR",
            },
        }

    @staticmethod
    def _generate_building_boq(
        materials: Dict[str, Any],
        costs: Dict[str, Any],
        rates: Dict[str, float],
    ) -> Dict[str, Any]:
        """Generate building-level BOQ with all sections."""
        sections = []
        sl_no = 0

        # ── Section 1: Earthwork ─────────────────────────────────────────
        earthwork_items = []
        excavation_vol = materials.get("excavation_volume_m3", 0)
        if excavation_vol > 0:
            sl_no += 1
            earthwork_items.append({
                "sl_no": sl_no,
                "description": "Excavation in ordinary soil for foundations",
                "unit": "m³",
                "quantity": round(excavation_vol, 2),
                "rate": rates.get("excavation_m3", 200),
                "amount": round(excavation_vol * rates.get("excavation_m3", 200)),
            })

        sections.append({
            "section_name": "A. Earthwork",
            "items": earthwork_items,
            "subtotal": sum(i["amount"] for i in earthwork_items),
        })

        # ── Section 2: RCC / Concrete ────────────────────────────────────
        concrete_items = []
        concrete_vol = materials.get("concrete_volume_m3", 0)
        if concrete_vol > 0:
            sl_no += 1
            concrete_items.append({
                "sl_no": sl_no,
                "description": "RCC M20 grade concrete (supply + placing + vibrating + curing)",
                "unit": "m³",
                "quantity": round(concrete_vol, 2),
                "rate": rates.get("concrete_rcc_m3", 5500),
                "amount": costs.get("concrete_cost", 0),
            })

        steel_kg = materials.get("steel_weight_kg", 0)
        if steel_kg > 0:
            sl_no += 1
            concrete_items.append({
                "sl_no": sl_no,
                "description": "Steel reinforcement Fe500D TMT bars (supply + bending + placing)",
                "unit": "kg",
                "quantity": round(steel_kg, 2),
                "rate": rates.get("steel_tmt_kg", 75),
                "amount": costs.get("steel_cost", 0),
            })

        sections.append({
            "section_name": "B. RCC & Structural Works",
            "items": concrete_items,
            "subtotal": sum(i["amount"] for i in concrete_items),
        })

        # ── Section 3: Masonry ───────────────────────────────────────────
        masonry_items = []
        bricks = materials.get("bricks_count", 0)
        blocks = materials.get("blocks_count", 0)

        if bricks > 0:
            sl_no += 1
            masonry_items.append({
                "sl_no": sl_no,
                "description": "Brick masonry in CM 1:6 with modular bricks",
                "unit": "nos",
                "quantity": bricks,
                "rate": rates.get("brick_red_pc", 10),
                "amount": costs.get("brick_cost", 0),
            })

        if blocks > 0:
            sl_no += 1
            masonry_items.append({
                "sl_no": sl_no,
                "description": "AAC block masonry 600×200×200mm with thin-bed mortar",
                "unit": "nos",
                "quantity": blocks,
                "rate": rates.get("block_aac_pc", 55),
                "amount": costs.get("block_cost", 0),
            })

        cement_bags = materials.get("cement_bags", 0)
        if cement_bags > 0:
            sl_no += 1
            masonry_items.append({
                "sl_no": sl_no,
                "description": "OPC 53 Grade Cement (50 kg bags)",
                "unit": "bags",
                "quantity": cement_bags,
                "rate": rates.get("cement_bag_50kg", 430),
                "amount": costs.get("cement_cost", 0),
            })

        sand_vol = materials.get("sand_volume_m3", 0)
        if sand_vol > 0:
            sl_no += 1
            masonry_items.append({
                "sl_no": sl_no,
                "description": "River sand / M-Sand (Zone II grading)",
                "unit": "m³",
                "quantity": round(sand_vol, 2),
                "rate": rates.get("sand_m3", 1400),
                "amount": costs.get("sand_cost", 0),
            })

        agg_vol = materials.get("aggregate_volume_m3", 0)
        if agg_vol > 0:
            sl_no += 1
            masonry_items.append({
                "sl_no": sl_no,
                "description": "Crushed stone aggregate 20mm nominal size",
                "unit": "m³",
                "quantity": round(agg_vol, 2),
                "rate": rates.get("aggregate_20mm_m3", 1600),
                "amount": costs.get("aggregate_cost", 0),
            })

        sections.append({
            "section_name": "C. Masonry & Materials",
            "items": masonry_items,
            "subtotal": sum(i["amount"] for i in masonry_items),
        })

        # ── Section 4: Finishing ─────────────────────────────────────────
        finishing_items = []

        plaster_area = materials.get("plaster_area_sq_m", 0)
        if plaster_area > 0:
            sl_no += 1
            finishing_items.append({
                "sl_no": sl_no,
                "description": "Cement plaster 12mm thick in CM 1:4 (internal walls & ceiling)",
                "unit": "m²",
                "quantity": round(plaster_area, 2),
                "rate": rates.get("plaster_m2", 280),
                "amount": costs.get("plaster_cost", 0),
            })

        paint_area = materials.get("paint_area_sq_m", 0)
        if paint_area > 0:
            sl_no += 1
            finishing_items.append({
                "sl_no": sl_no,
                "description": "Interior emulsion paint — 2 coats over 1 coat primer",
                "unit": "m²",
                "quantity": round(paint_area, 2),
                "rate": rates.get("paint_interior_m2", 120),
                "amount": costs.get("paint_cost", 0),
            })

        tile_area = materials.get("tile_area_sq_m", 0)
        if tile_area > 0:
            sl_no += 1
            finishing_items.append({
                "sl_no": sl_no,
                "description": "Vitrified floor tiles 600×600mm (supply + fixing with tile adhesive)",
                "unit": "m²",
                "quantity": round(tile_area, 2),
                "rate": rates.get("tiles_m2", 650),
                "amount": costs.get("tiles_cost", 0),
            })

        wp_area = materials.get("waterproofing_area_sq_m", 0)
        if wp_area > 0:
            sl_no += 1
            finishing_items.append({
                "sl_no": sl_no,
                "description": "Waterproofing with liquid membrane — 2 coats (roof + wet areas)",
                "unit": "m²",
                "quantity": round(wp_area, 2),
                "rate": rates.get("waterproofing_m2", 380),
                "amount": costs.get("waterproofing_cost", 0),
            })

        sections.append({
            "section_name": "D. Finishing Works",
            "items": finishing_items,
            "subtotal": sum(i["amount"] for i in finishing_items),
        })

        # ── Section 5: MEP (Plumbing + Electrical) ───────────────────────
        mep_items = []

        plumbing_cost = costs.get("plumbing_cost", 0)
        if plumbing_cost > 0:
            sl_no += 1
            mep_items.append({
                "sl_no": sl_no,
                "description": "Plumbing works (pipes, fittings, fixtures, infrastructure)",
                "unit": "LS",
                "quantity": 1,
                "rate": plumbing_cost,
                "amount": plumbing_cost,
            })

        electrical_cost = costs.get("electrical_cost", 0)
        if electrical_cost > 0:
            sl_no += 1
            mep_items.append({
                "sl_no": sl_no,
                "description": "Electrical works (wiring, switches, lights, fans, infrastructure)",
                "unit": "LS",
                "quantity": 1,
                "rate": electrical_cost,
                "amount": electrical_cost,
            })

        door_cost = costs.get("door_cost", 0)
        if door_cost > 0:
            sl_no += 1
            mep_items.append({
                "sl_no": sl_no,
                "description": "Doors (main + flush + PVC with frames and hardware)",
                "unit": "LS",
                "quantity": 1,
                "rate": door_cost,
                "amount": door_cost,
            })

        window_cost = costs.get("window_cost", 0)
        if window_cost > 0:
            sl_no += 1
            mep_items.append({
                "sl_no": sl_no,
                "description": "Aluminium sliding windows with glass panels",
                "unit": "LS",
                "quantity": 1,
                "rate": window_cost,
                "amount": window_cost,
            })

        sections.append({
            "section_name": "E. MEP, Doors & Windows",
            "items": mep_items,
            "subtotal": sum(i["amount"] for i in mep_items),
        })

        # ── Grand Total ──────────────────────────────────────────────────
        material_total = sum(s["subtotal"] for s in sections)

        return {
            "sections": sections,
            "material_subtotal": material_total,
            "labour_cost": costs.get("labour_cost", 0),
            "equipment_cost": costs.get("equipment_cost", 0),
            "contractor_margin": costs.get("contractor_margin", 0),
            "contingency": costs.get("contingency", 0),
            "transport_cost": costs.get("transport_cost", 0),
            "gst_amount": costs.get("gst_amount", 0),
            "miscellaneous": costs.get("miscellaneous", 0),
            "grand_total": costs.get("grand_total", 0),
        }

    @staticmethod
    def _generate_room_wise_boq(
        rooms: List[Dict[str, Any]],
        floor_height_m: float,
        wall_thickness_m: float,
        brick_type: str,
        rates: Dict[str, float],
    ) -> List[Dict[str, Any]]:
        """Generate per-room material BOQ."""
        room_boqs = []

        for room in rooms:
            label = room.get("label", "Room")
            area_m2 = room.get("area_m2", 0)
            perimeter_m = room.get("perimeter_m", 0)

            if area_m2 <= 0:
                continue

            # Calculate room-specific materials
            wall_area_m2 = perimeter_m * floor_height_m if perimeter_m > 0 else 0
            wall_volume_m3 = wall_area_m2 * wall_thickness_m * 0.9  # 10% openings
            ceiling_area = area_m2
            floor_area = area_m2

            # Bricks/blocks
            units_per_m3 = BRICKS_PER_M3 if "aac" not in brick_type else AAC_BLOCKS_PER_M3
            masonry_count = math.ceil(wall_volume_m3 * units_per_m3 * 1.05)

            # Mortar
            mortar_vol = wall_volume_m3 * MORTAR_FRACTION_OF_MASONRY * MORTAR_DRY_VOLUME_FACTOR
            cement_bags = math.ceil((mortar_vol / 7) / CEMENT_BAG_VOLUME_M3)

            # Plaster
            plaster_area = wall_area_m2 * 2 + ceiling_area  # Both sides + ceiling
            plaster_vol = plaster_area * (PLASTER_THICKNESS_MM["internal"] / 1000)

            # Paint
            paint_area = plaster_area * 1.2

            # Tiles
            tile_area = floor_area

            items = []
            sl = 0

            if masonry_count > 0:
                sl += 1
                unit_rate = rates.get("brick_red_pc", 10) if "aac" not in brick_type else rates.get("block_aac_pc", 55)
                items.append({
                    "sl_no": sl,
                    "description": "Masonry units (bricks/blocks)",
                    "unit": "nos",
                    "quantity": masonry_count,
                    "rate": unit_rate,
                    "amount": round(masonry_count * unit_rate),
                })

            if cement_bags > 0:
                sl += 1
                items.append({
                    "sl_no": sl,
                    "description": "Cement (50kg bags)",
                    "unit": "bags",
                    "quantity": cement_bags,
                    "rate": rates.get("cement_bag_50kg", 430),
                    "amount": round(cement_bags * rates.get("cement_bag_50kg", 430)),
                })

            if plaster_area > 0:
                sl += 1
                items.append({
                    "sl_no": sl,
                    "description": "Plaster (12mm CM 1:4)",
                    "unit": "m²",
                    "quantity": round(plaster_area, 2),
                    "rate": rates.get("plaster_m2", 280),
                    "amount": round(plaster_area * rates.get("plaster_m2", 280)),
                })

            if paint_area > 0:
                sl += 1
                items.append({
                    "sl_no": sl,
                    "description": "Interior emulsion paint (2 coats)",
                    "unit": "m²",
                    "quantity": round(paint_area, 2),
                    "rate": rates.get("paint_interior_m2", 120),
                    "amount": round(paint_area * rates.get("paint_interior_m2", 120)),
                })

            if tile_area > 0:
                sl += 1
                items.append({
                    "sl_no": sl,
                    "description": "Floor tiles (vitrified 600×600mm)",
                    "unit": "m²",
                    "quantity": round(tile_area, 2),
                    "rate": rates.get("tiles_m2", 650),
                    "amount": round(tile_area * rates.get("tiles_m2", 650)),
                })

            # Room-type specific additions
            label_lower = label.lower()
            if "bath" in label_lower or "toilet" in label_lower:
                sl += 1
                wp_area = area_m2 + wall_area_m2 * 0.5  # Floor + half wall height
                items.append({
                    "sl_no": sl,
                    "description": "Waterproofing (liquid membrane)",
                    "unit": "m²",
                    "quantity": round(wp_area, 2),
                    "rate": rates.get("waterproofing_m2", 380),
                    "amount": round(wp_area * rates.get("waterproofing_m2", 380)),
                })

            subtotal = sum(i["amount"] for i in items)

            room_boqs.append({
                "room_id": room.get("id", ""),
                "room_name": label,
                "area_m2": round(area_m2, 2),
                "area_sqft": round(area_m2 * 10.764, 2),
                "items": items,
                "subtotal": subtotal,
            })

        return room_boqs

    # ══════════════════════════════════════════════════════════════════════
    # EXPORT METHODS
    # ══════════════════════════════════════════════════════════════════════

    @staticmethod
    def export_csv(boq_data: Dict[str, Any]) -> str:
        """Export BOQ to CSV string."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(["BuildWise AI — Bill of Quantities"])
        writer.writerow(["Project:", boq_data.get("project_name", "")])
        writer.writerow(["Generated:", boq_data.get("generated_at", "")])
        writer.writerow([])

        # Building BOQ
        building = boq_data.get("building_boq", {})
        for section in building.get("sections", []):
            writer.writerow([section["section_name"]])
            writer.writerow(["Sl.No", "Description", "Unit", "Quantity", "Rate (₹)", "Amount (₹)"])

            for item in section.get("items", []):
                writer.writerow([
                    item["sl_no"],
                    item["description"],
                    item["unit"],
                    item["quantity"],
                    item["rate"],
                    item["amount"],
                ])

            writer.writerow(["", "", "", "", "Subtotal:", section["subtotal"]])
            writer.writerow([])

        # Totals
        writer.writerow(["", "", "", "", "Material Subtotal:", building.get("material_subtotal", 0)])
        writer.writerow(["", "", "", "", "Labour Cost:", building.get("labour_cost", 0)])
        writer.writerow(["", "", "", "", "Equipment Cost:", building.get("equipment_cost", 0)])
        writer.writerow(["", "", "", "", "Contractor Margin:", building.get("contractor_margin", 0)])
        writer.writerow(["", "", "", "", "GST (18%):", building.get("gst_amount", 0)])
        writer.writerow(["", "", "", "", "GRAND TOTAL:", building.get("grand_total", 0)])

        # Room-wise BOQ
        writer.writerow([])
        writer.writerow(["ROOM-WISE BREAKDOWN"])
        for room_boq in boq_data.get("room_wise_boq", []):
            writer.writerow([])
            writer.writerow([f"{room_boq['room_name']} ({room_boq['area_sqft']} sq.ft)"])
            writer.writerow(["Sl.No", "Description", "Unit", "Quantity", "Rate (₹)", "Amount (₹)"])
            for item in room_boq.get("items", []):
                writer.writerow([
                    item["sl_no"], item["description"], item["unit"],
                    item["quantity"], item["rate"], item["amount"],
                ])
            writer.writerow(["", "", "", "", "Room Total:", room_boq["subtotal"]])

        return output.getvalue()

    @staticmethod
    def export_pdf(boq_data: Dict[str, Any]) -> bytes:
        """Export BOQ to PDF using ReportLab."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=20*mm, bottomMargin=15*mm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('BOQTitle', parent=styles['Heading1'], fontSize=16,
                                     textColor=colors.HexColor('#7C3AED'), spaceAfter=8)
        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12,
                                       textColor=colors.HexColor('#1E1E2E'), spaceAfter=4)
        normal_style = styles['Normal']

        elements = []

        # Title
        elements.append(Paragraph("BuildWise AI — Bill of Quantities", title_style))
        elements.append(Paragraph(f"Project: {boq_data.get('project_name', 'N/A')}", normal_style))
        elements.append(Paragraph(f"Generated: {boq_data.get('generated_at', 'N/A')}", normal_style))
        elements.append(Spacer(1, 10*mm))

        building = boq_data.get("building_boq", {})

        for section in building.get("sections", []):
            elements.append(Paragraph(section["section_name"], section_style))

            table_data = [["Sl.", "Description", "Unit", "Qty", "Rate (₹)", "Amount (₹)"]]
            for item in section.get("items", []):
                table_data.append([
                    str(item["sl_no"]),
                    item["description"][:60],
                    item["unit"],
                    str(item["quantity"]),
                    f"{item['rate']:,.0f}",
                    f"{item['amount']:,.0f}",
                ])

            table_data.append(["", "", "", "", "Subtotal:", f"₹{section['subtotal']:,.0f}"])

            t = Table(table_data, colWidths=[20*mm, 75*mm, 15*mm, 20*mm, 22*mm, 28*mm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7C3AED')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F5F0FF')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#FAFAFA')]),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 5*mm))

        # Grand total section
        totals_data = [
            ["Material Subtotal", f"₹{building.get('material_subtotal', 0):,.0f}"],
            ["Labour Cost", f"₹{building.get('labour_cost', 0):,.0f}"],
            ["Equipment Cost", f"₹{building.get('equipment_cost', 0):,.0f}"],
            ["Contractor Margin", f"₹{building.get('contractor_margin', 0):,.0f}"],
            ["GST (18%)", f"₹{building.get('gst_amount', 0):,.0f}"],
            ["GRAND TOTAL", f"₹{building.get('grand_total', 0):,.0f}"],
        ]
        tt = Table(totals_data, colWidths=[100*mm, 60*mm])
        tt.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#7C3AED')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(Spacer(1, 8*mm))
        elements.append(tt)

        doc.build(elements)
        return buffer.getvalue()

    @staticmethod
    def export_excel(boq_data: Dict[str, Any]) -> bytes:
        """Export BOQ to Excel using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")

        wb = Workbook()

        # ── Sheet 1: Building BOQ ────────────────────────────────────────
        ws = wb.active
        ws.title = "Building BOQ"

        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        title_font = Font(name='Calibri', bold=True, size=14, color='7C3AED')

        ws.merge_cells('A1:F1')
        ws['A1'] = "BuildWise AI — Bill of Quantities"
        ws['A1'].font = title_font

        ws['A2'] = f"Project: {boq_data.get('project_name', '')}"
        ws['A3'] = f"Generated: {boq_data.get('generated_at', '')}"

        row = 5
        building = boq_data.get("building_boq", {})

        for section in building.get("sections", []):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1, value=section["section_name"]).font = Font(bold=True, size=11)
            row += 1

            headers = ["Sl.No", "Description", "Unit", "Quantity", "Rate (₹)", "Amount (₹)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            row += 1

            for item in section.get("items", []):
                ws.cell(row=row, column=1, value=item["sl_no"])
                ws.cell(row=row, column=2, value=item["description"])
                ws.cell(row=row, column=3, value=item["unit"])
                ws.cell(row=row, column=4, value=item["quantity"])
                ws.cell(row=row, column=5, value=item["rate"])
                ws.cell(row=row, column=6, value=item["amount"])
                row += 1

            ws.cell(row=row, column=5, value="Subtotal:").font = Font(bold=True)
            ws.cell(row=row, column=6, value=section["subtotal"]).font = Font(bold=True)
            row += 2

        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18

        # ── Sheet 2: Room-wise BOQ ───────────────────────────────────────
        ws2 = wb.create_sheet("Room-wise BOQ")

        ws2.merge_cells('A1:F1')
        ws2['A1'] = "Room-wise Material Breakdown"
        ws2['A1'].font = title_font

        row = 3
        for room_boq in boq_data.get("room_wise_boq", []):
            ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws2.cell(row=row, column=1,
                     value=f"{room_boq['room_name']} — {room_boq['area_sqft']} sq.ft").font = Font(bold=True, size=11)
            row += 1

            headers = ["Sl.No", "Description", "Unit", "Quantity", "Rate (₹)", "Amount (₹)"]
            for col, h in enumerate(headers, 1):
                cell = ws2.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
            row += 1

            for item in room_boq.get("items", []):
                ws2.cell(row=row, column=1, value=item["sl_no"])
                ws2.cell(row=row, column=2, value=item["description"])
                ws2.cell(row=row, column=3, value=item["unit"])
                ws2.cell(row=row, column=4, value=item["quantity"])
                ws2.cell(row=row, column=5, value=item["rate"])
                ws2.cell(row=row, column=6, value=item["amount"])
                row += 1

            ws2.cell(row=row, column=5, value="Room Total:").font = Font(bold=True)
            ws2.cell(row=row, column=6, value=room_boq["subtotal"]).font = Font(bold=True)
            row += 2

        ws2.column_dimensions['A'].width = 8
        ws2.column_dimensions['B'].width = 45
        ws2.column_dimensions['C'].width = 10
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 15
        ws2.column_dimensions['F'].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
